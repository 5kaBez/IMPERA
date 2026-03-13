#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys

# Set UTF-8 encoding for console output on Windows
os.environ['PYTHONIOENCODING'] = 'utf-8'

import pandas as pd
import openpyxl
import re
import os
import sys


def normalize_name(name: str) -> str:
    """
    Унифицировать названия: привести к Title Case и убрать лишние пробелы.
    Преобразует "МЕНЕДЖМЕНТ", "менеджмент", "М Е Н Е Д Ж М Е Н Т" в "Менеджмент"
    
    Алгоритм:
    1. Убираем лишние пробелы
    2. Если это буквы разделённые пробелом (М Е Н) - склеиваем их
    3. Приводим в нижний регистр
    4. Заменяем Ё на Е для унификации (Учёт → Учет)
    5. Делаем first letter upper для каждого слова
    """
    if not name or name == '-':
        return name
    
    # Убираем лишние пробелы (включая переносы строк)
    name = ' '.join(name.split())
    
    # Убираем пробелы между отдельными буквами (М Е Н Е Д Ж М Е Н Т)
    words = name.split()
    if all(len(word) == 1 and word.isalpha() for word in words):
        name = ''.join(words)
    
    # Приводим в нижний регистр для консистентности
    name = name.lower()
    
    # Заменяем Ё на Е для унификации (Учёт → учет → Учет)
    name = name.replace('ё', 'е')
    
    # Делаем первую букву каждого слова прописной
    # Но сохраняем дефисы и другие разделители
    result = []
    capitalize_next = True
    for char in name:
        if char in ' -':
            result.append(char)
            capitalize_next = True
        elif capitalize_next and char.isalpha():
            result.append(char.upper())
            capitalize_next = False
        else:
            result.append(char)
    
    return ''.join(result)


def extract_institute_from_program(program_name: str) -> str:
    """
    Извлечь название организации из названия программы.
    Используется для ОЗФО и магистратуры, где название организации
    указано в описании программы.
    
    Порядок проверки важен! Проверяем специфичные сначала, общие - потом.
    """
    if not program_name or program_name == '-':
        return ""
    
    program_upper = program_name.upper()
    
    # ИНФОРМАЦИОННЫЕ СИСТЕМЫ (проверяется ПЕРВЫМ, т.к. может быть 'бизнес-информатика')
    if any(kw in program_upper for kw in ['ПРОГРАММ', 'КОМПЬЮТЕР', 'ДАННЫХ', 'СЕТЕВЫЕ', 'СЕТИ']):
        return 'Институт информационных систем'
    # Специфичная проверка для информатики (но БЕЗ бизнеса)
    if 'ИНФОРМАТИКА' in program_upper and 'БИЗНЕС' not in program_upper:
        return 'Институт информационных систем'
    # Если ИНФОРМАТИК/ИНФОРМАЦ есть (включая БИЗНЕС-ИНФОРМАТИКА)
    if any(kw in program_upper for kw in ['ИНФОРМАТИК', 'ИНФОРМАЦ']):
        return 'Институт информационных систем'
    
    # ЭКОНОМИКА И ФИНАНСЫ (учет, аудит, бухгалтерия, экономика, финансы, налоги)
    if any(kw in program_upper for kw in ['БУХГАЛТЕР', 'УЧЕТ', 'АУДИТ', 'ЭКОНОМИ', 'ФИНАНС', 'НАЛОГ', 'СТАТИСТИ']):
        return 'Институт экономики и финансов'
    
    # ГОСУДАРСТВЕННОЕ УПРАВЛЕНИЕ И ПРАВО
    if any(kw in program_upper for kw in ['ГОСУДАРСТВЕН', 'МУНИЦИПАЛЬ', 'ПРАВО', 'ЮРИД', 'БЕЗОПАСНОСТ', 'ТАМОЖЕН']):
        return 'Институт государственного управления и права'
    
    # УПРАВЛЕНИЕ ПЕРСОНАЛОМ, СОЦИАЛЬНЫЕ, БИЗНЕС-КОММУНИКАЦИИ (проверяем конкретные слова)
    if any(kw in program_upper for kw in ['ПЕРСОНАЛ', 'СОЦИАЛЬ', 'КОММУНИКАЦ', 'СВЯЗИ', 'ОБРАЗОВАТ', 'ПСИХОЛОГ', 'ТУРИЗМ']):
        return 'Институт управления персоналом, социальных и бизнес-коммуникаций'
    
    # БИЗНЕС (только если нет ИНФОРМАТИКИ)
    if 'БИЗНЕС' in program_upper and 'ИНФОРМАТИКА' not in program_upper:
        return 'Институт управления персоналом, социальных и бизнес-коммуникаций'
    
    # МАРКЕТИНГ  
    if any(kw in program_upper for kw in ['МАРКЕТИНГ', 'РЕКЛАМ', 'ТОРГОВЛ', 'РОЗНИЦА', 'ПРОДАЖ']):
        return 'Институт маркетинга'
    
    # МЕНЕДЖМЕНТ и ПРОЕКТНОЕ УПРАВЛЕНИЕ
    if any(kw in program_upper for kw in ['МЕНЕДЖМЕНТ', 'ПРОИЗВОД', 'ЛОГИСТИК', 'КАЧЕСТВО', 'ОПЕРАЦИОН', 'СТРАТЕГИ', 'ЭНЕРГЕТИ', 'ПРОЕКТ']):
        return 'Институт отраслевого менеджмента'
    
    return ""


def get_institute_name(sheet_name: str, program_name: str = '', direction_name: str = '', determined_institute: str = '') -> str:
    """
    Определить название института по названию листа, программе и направлению.
    
    Приоритет:
    1. Если передан determined_institute (из объединённых ячеек метаданных) - используем его
    2. Иначе используем логику определения по названиям
    
    Для листов с ДВА Instituto (например "ИЭФ, ИГУИП"), используем направление для разделения:
    - ИЭФ направления: ЭКОНОМИКА, ФИНАНСЫ И КРЕДИТ
    - ИГУИП направления: ЮРИСПРУДЕНЦИЯ, ГМУ, ПОЛИТОЛОГИЯ, СОЦИОЛОГИЯ
    - ИОМ направления: МЕНЕДЖМЕНТ, ИННОВАТИКА
    - ИМ направления: МЕНЕДЖМЕНТ, РЕКЛАМА
    """
    # Если уже определен институт из объединённых ячеек - используем его!
    if determined_institute and determined_institute.strip():
        return determined_institute
    
    sheet_lower = sheet_name.lower()
    program_upper = program_name.upper()
    direction_upper = direction_name.upper()
    
    if 'игуип' in sheet_lower:
        # На этом листе может быть только ИГУИП, но проверим направление
        if 'ЭКОНОМИК' in direction_upper or 'ФИНАНС' in direction_upper:
            return 'Институт экономики и финансов'  # На случай ошибки в структуре
        return 'Институт государственного управления и права'
    elif 'иэф' in sheet_lower or 'иэи' in sheet_lower:
        # ТОЛЬКО ИЭФ, не смешано
        return 'Институт экономики и финансов'
    elif 'иом' in sheet_lower and 'им' in sheet_lower:
        # ДВА INSTITUTE НА ОДНОМ ЛИСТЕ: "1 к. маг ИОМ, ИМ"
        # Различаем по направлению или программе
        if 'ИННОВАТИК' in direction_upper:
            return 'Институт отраслевого менеджмента'
        elif 'МАРКЕТИНГ' in direction_upper or 'РЕКЛАМ' in direction_upper:
            return 'Институт маркетинга'
        elif 'МАРКЕТИНГ' in program_upper or 'РЕКЛАМ' in program_upper or 'ЦИФРОВОЙ МАРКЕТИНГ' in program_upper:
            return 'Институт маркетинга'
        elif 'ЭНЕРГЕТ' in program_upper or 'ИНВЕСТИЦ' in program_upper or 'ПРОЕКТ' in program_upper:
            return 'Институт отраслевого менеджмента'
        else:
            # По-умолчанию для менеджмента-подобных
            return 'Институт отраслевого менеджмента'
    elif 'иом' in sheet_lower:
        return 'Институт отраслевого менеджмента'
    elif 'им' in sheet_lower and 'иупс' not in sheet_lower:
        return 'Институт маркетинга'
    elif 'иупс' in sheet_lower or 'иупсибк' in sheet_lower:
        # На этом листе огут быть две разные организации - различаем по программе
        if 'ИНФОРМАЦИОНН' in program_upper or ('СИСТЕМЫ' in program_upper and 'ИНФОРМАЦ' in program_upper):
            return 'Институт информационных систем'
        else:
            return 'Институт управления персоналом, социальных и бизнес-коммуникаций'
    elif 'иис' in sheet_lower:
        return 'Институт информационных систем'
    else:
        return sheet_name


def normalize_magistracy_direction(direction: str, program: str) -> tuple:
    """
    Унифицирует названия направлений магистратуры ПО НАЗВАНИЮ ПРОГРАММЫ.
    Возвращает кортеж (normalized_direction, normalized_program).
    
    ВАЖНО: Если программа пуста, используем направление как программу.
    """
    if not direction or direction == '-':
        return direction, program
    if not program or program == '-':
        program = ''
    
    # Очищаем от лишних пробелов
    direction = ' '.join(direction.split())
    program = ' '.join(program.split()) if program else ''
    
    # Если программа пуста, используем направление как программу
    if not program:
        program = direction
        # DEBUG: для магистратуры, если программа пустая
        # print(f"DEBUG: Empty program detected! Direction: {direction}, Program set to: {program}")
    
    # Если в слове все буквы разделены пробелами (М Е Н Е Д Ж), склеиваем их
    if all(len(word) == 1 for word in direction.split()):
        direction = ''.join(direction.split())
    
    direction_upper = direction.upper()
    program_upper = program.upper()
    
    # === ОПРЕДЕЛЯЕМ НАПРАВЛЕНИЕ ПО ПРОГРАММЕ (приоритет) ===
    
    # ГОСТИНИЧНОЕ ДЕЛО
    if any(kw in program_upper for kw in ['ГОСТИН', 'ИВЕНТ']):
        return normalize_name('Гостиничное дело'), normalize_name(program)
    
    # ИННОВАТИКА
    if any(kw in program_upper for kw in ['УСТОЙЧИВЫМ', 'УСТОЙЧИВО']):
        return normalize_name('Инноватика'), normalize_name(program)
    
    # ФИНАНСЫ И КРЕДИТ - программы ИЛИ направление
    if any(kw in program_upper for kw in ['ФИНАНСОВЫЕ РЫНКИ', 'ФОНДОВ', 'БАНКИНГ', 'КРЕДИТ']) or \
       any(kw in direction_upper for kw in ['ФИНАНС']):
        return normalize_name('Финансы и кредит'), normalize_name(program)
    
    # ЭКОНОМИКА - ПРИОРИТЕТ (перед БИЗНЕС-ИНФОРМАТИКОЙ, перед МЕНЕДЖМЕНТОМ)
    if any(kw in program_upper for kw in [
        'ПРИКЛАДНОЙ АНАЛИЗ', 'ЭКОНОМИКА ИНТЕГРАЦИОН', 'МЕЖДУНАРОДНАЯ ЭКОНОМИКА',
        'ИНФОРМАЦИОННО-АНАЛИТИЧЕ', 'АНАЛИТИЧЕСКИЕ ТЕХНОЛОГИИ',
        'УПРАВЛЕНЧЕСКАЯ ЭКОНОМИКА'
    ]):
        return normalize_name('Экономика'), normalize_name(program)
    
    # БИЗНЕС-ИНФОРМАТИКА
    if any(kw in program_upper for kw in ['ЦИФРОВОЙ ТРАНСФОРМ', 'ИНФОРМАТИК', 'ИНФОРМАЦ']):
        if 'ТРАНСФОРМ' in program_upper or 'ЦИФРОВ' in program_upper:
            return normalize_name('Бизнес-Информатика'), normalize_name(program)
    
    # РЕКЛАМА И СВЯЗИ С ОБЩЕСТВЕННОСТЬЮ
    if any(kw in program_upper for kw in ['ПРОДЮСИРОВАНИЕ РЕКЛ', 'РЕКЛАМНЫХ КОММУН']):
        return normalize_name('Реклама И Связи С Общественностью'), normalize_name(program)
    
    # УПРАВЛЕНИЕ ПЕРСОНАЛОМ
    if 'УПРАВЛЕНИЕ ПЕРСОНАЛОМ' in program_upper:
        return normalize_name('Управление Персоналом'), normalize_name(program)
    
    # ГОСУДАРСТВЕННОЕ И МУНИЦИПАЛЬНОЕ УПРАВЛЕНИЕ
    if any(kw in program_upper for kw in ['ГОСУДАРСТВЕННОЕ И МУНИЦИПАЛЬНОЕ', 'ГОСУДАРСТВ И МУНИЦ']):
        return normalize_name('Государственное И Муниципальное Управление'), normalize_name(program)
    
    # ЮРИСПРУДЕНЦИЯ
    if any(kw in program_upper for kw in ['ПРАВОВОЕ ОБЕСПЕЧЕНИЕ', 'ЦИФРОВОЕ ПРАВО']):
        return normalize_name('Юриспруденция'), normalize_name(program)
    
    # ПОЛИТОЛОГИЯ
    if any(kw in program_upper for kw in ['ПОЛИТИЧЕСКИЙ МЕНЕДЖМЕНТ', 'ПОЛИТИЧЕСК']):
        return normalize_name('Политология'), normalize_name(program)
    
    # === МЕНЕДЖМЕНТ (по-умолчанию для многих программ и направлений) ===
    # Включает: маркетинг, киберспорт, спорт, творчество, проекты, энергетику, производство и др.
    if any(kw in program_upper for kw in [
        'МАРКЕТИНГ', 'КИБЕРСПОРТ', 'ФИДЖИ', 'СПОРТИВ', 'ФИТНЕС',
        'ТВОРЧЕСК', 'ПРОЕКТИРОВАНИЕ', 'КАПИТАЛИЗАЦИЯ', 'ПРОЕКТ',
        'СТРАТЕГИЧЕСКОЕ', 'ЦИФРОВОЙ МАРКЕТИНГ', 'ЭНЕРГЕТИЧЕСКИЙ', 
        'ИНВЕСТИЦИОННО', 'ПРОИЗВОДСТВЕНН', 'ОРГАНИЗАТОР', 
        'УПРАВЛЕНИЕ МЕЖДУНАРОДНЫМ', 'УПРАВЛЕНИЕ ПРОЕКТ'
    ]):
        return normalize_name('Менеджмент'), normalize_name(program)
    
    # === Fallback по НАЗВАНИЮ НАПРАВЛЕНИЯ (для частичной совместимости) ===
    
    if any(kw in direction_upper for kw in ['ОТРАСЛЕВ', 'МАРКЕТИНГ']):
        return normalize_name('Менеджмент'), normalize_name(program)
    
    if 'ИНФОРМАЦ' in direction_upper:
        return normalize_name('Бизнес-Информатика'), normalize_name(program)
    
    if 'ЭКОНОМИК' in direction_upper:
        return normalize_name('Экономика'), normalize_name(program)
    
    if 'ФИНАНС' in direction_upper:
        return normalize_name('Финансы И Кредит'), normalize_name(program)
    
    if any(kw in direction_upper for kw in ['ГОСУДАРСТВЕН', 'ПРАВО', 'ЮРИСПРУД']):
        return normalize_name('Юриспруденция'), normalize_name(program)
    
    if 'ПОЛИТОЛОГ' in direction_upper:
        return normalize_name('Политология'), normalize_name(program)
    
    if 'СОЦИОЛОГ' in direction_upper:
        return normalize_name('Социология'), normalize_name(program)
    
    # Catch-all для ВЕРХНИХ РЕГИСТРОВ
    if direction_upper == direction:
        return normalize_name(direction), normalize_name(program)
    
    # Нормализуем перед возвращением
    return normalize_name(direction), normalize_name(program)


def find_header_row(ws):
    """
    Найти строку с заголовками групп.
    Ищет '№ учебной группы' (основной формат),
    если не найдено - ищет строку с 'День' (магистратура/заочка).
    """
    # Первый вариант: ищем "№ учебной группы"
    for i in range(1, min(20, ws.max_row)):
        for j in range(1, ws.max_column):
            cell = ws.cell(i, j)
            if cell.value and '№ учебной группы' in str(cell.value):
                return i
    
    # Второй вариант: для магистратуры/заочки ищем строку, предшествующую "День"
    # Обычно это строка с номерами групп/подгрупп
    for i in range(1, min(20, ws.max_row)):
        for j in range(1, ws.max_column):
            cell = ws.cell(i, j)
            if cell.value and 'День' in str(cell.value):
                # Заголовок с "День" найден, возвращаем эту строку как header
                # (для магистратуры/заочки это строка 8, где и День, и группы)
                return i
    
    return None


def find_time_row(ws, header_row):
    """
    Найти строку с днями недели/временем.
    
    Для основного формата (1-4.xlsx): header_row содержит "№ учебной группы",
    time_row находится ниже (содержит "День").
    
    Для магистратуры/заочки: header_row уже содержит "День",
    поэтому time_row == header_row.
    """
    if header_row is None:
        return None
    
    # Проверим, содержит ли сам header_row "День"
    for j in range(1, ws.max_column):
        cell = ws.cell(header_row, j)
        if cell.value and 'День' in str(cell.value):
            # header_row - это уже строка с "День"
            return header_row
    
    # Иначе ищем "День" ниже header_row
    for i in range(header_row + 1, min(header_row + 5, ws.max_row)):
        for j in range(1, min(5, ws.max_column)):
            cell = ws.cell(i, j)
            if cell.value and 'День' in str(cell.value):
                return i
    
    # По умолчанию - строка ниже header
    if header_row + 1 <= ws.max_row:
        return header_row + 1
    return None


def get_group_columns(ws, header_row):
    """
    Получить номера колонок с номерами групп.
    
    Для основного формата: ищет числовые значения в header_row.
    Для магистратуры/заочки: header_row содержит "День",
    поэтому ищет числовые значения в строке ВЫШЕ.
    """
    groups_dict = {}
    
    if header_row is None:
        return groups_dict
    
    # Сначала ищем в самом header_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        if cell.value:
            try:
                group_num = int(cell.value)
                groups_dict[col] = group_num
            except (ValueError, TypeError):
                pass
    
    # Если в header_row ничего не нашлось, попробуем строку выше
    # (это для магистратуры/заочки, где header_row содержит "День")
    if not groups_dict and header_row > 1:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(header_row - 1, col)
            if cell.value:
                try:
                    group_num = int(cell.value)
                    groups_dict[col] = group_num
                except (ValueError, TypeError):
                    pass
    
    return groups_dict


def get_group_metadata(ws, header_row, group_cols, is_magistracy=False, sheet_name: str = ''):
    """Получить направление, программу и институт для каждой группы."""
    metadata = {}
    
    # Ищем строки с НАПРАВЛЕНИЕМ и ПРОГРАММОЙ
    direction_row = None
    program_row = None
    
    for i in range(1, header_row):
        for j in range(1, 5):
            cell_val = ws.cell(i, j)
            if cell_val.value:
                val_str = str(cell_val.value).upper()
                if 'НАПРАВЛЕНИЕ' in val_str:
                    direction_row = i
                elif 'ОБРАЗОВАТЕЛЬНАЯ' in val_str or 'ПРОГРАММА' in val_str:
                    program_row = i
    
    # Определяем границы каждого института на основе объединённых ячеек в строке 4
    # ТОЛЬКО для листа с ИУПСИБК, ИИС!
    institute_ranges = {}  # {col_index: institute_name}
    
    sheet_lower = sheet_name.lower()
    if 'иупс' in sheet_lower or 'иупсибк' in sheet_lower:
        # Специальная логика для листа с ИУПСИБК/ИИС
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 4 and merged_range.max_row == 4:
                # Это объединённая ячейка в строке ИНСТИТУТА
                cell = ws.cell(4, merged_range.min_col)
                institute_text = str(cell.value).strip() if cell.value else ""
                # Очищаем от лишних пробелов
                institute_text = ' '.join(institute_text.split()).upper()
                
                # Определяем название института
                if 'ИНФОРМАЦИОННЫХ СИСТЕМ' in institute_text:
                    institute_name = 'Институт информационных систем'
                elif 'УПРАВЛЕНИЯ ПЕРСОНАЛОМ' in institute_text or 'СОЦИАЛЬНЫХ' in institute_text:
                    institute_name = 'Институт управления персоналом, социальных и бизнес-коммуникаций'
                else:
                    # Пропускаем ячейки, которые просто содержат "ИНСТИТУТ"
                    continue
                
                # Добавляем все колонки в этом диапазоне
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    institute_ranges[col] = institute_name
    
    # Извлекаем данные для каждой колонки с группой
    last_direction = '-'
    for col_idx in sorted(group_cols.keys()):
        direction = '-'
        program = '-'
        institute = ''
        
        # Получаем институт если определено объединёнными ячејками (ТОЛЬКО для ИУПСИБК)
        if col_idx in institute_ranges:
            institute = institute_ranges[col_idx]
        
        # Получаем направление с "fill forward" логикой
        if direction_row:
            cell = ws.cell(direction_row, col_idx)
            if cell.value and str(cell.value).strip() != '-':
                direction = normalize_name(str(cell.value).strip())
                last_direction = direction
            else:
                direction = last_direction
        
        # Получаем программу
        if program_row:
            cell = ws.cell(program_row, col_idx)
            if cell.value:
                program = normalize_name(str(cell.value).strip())
        
        # Применяем нормализацию для магистратуры
        if is_magistracy and direction != '-':
            direction, program = normalize_magistracy_direction(direction, program)
        
        metadata[col_idx] = {
            'direction': direction,
            'program': program,
            'institute': institute,  # Добавляем информацию о институте (может быть пуста)
        }
    
    return metadata


def get_day_time_columns(ws, time_row):
    """Получить номера колонок для Дня, Времени, Недели."""
    day_col = None
    time_col = None
    week_col = None
    
    if time_row is None:
        return day_col, time_col, week_col
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(time_row, col)
        if not cell.value:
            continue
        
        val = str(cell.value).strip().lower()
        if 'день' in val:
            day_col = col
        elif 'время' in val:
            time_col = col
        elif 'неделя' in val or 'четн' in val:  # Ищем 'четн' вместо 'четность'
            week_col = col
    
    return day_col, time_col, week_col


def is_cell_merged(ws, row, col):
    """Проверить объединена ли ячейка."""
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return True
    return False


def parse_cell_content(cell_text: str):
    """
    Парсить содержимое ячейки на component'ы.
    
    Может быть в одной строке:
    "Предмет (Л 1-12н) Преподаватель А.Б. Аудитория"
    
    Или в разных строках.
    """
    lines = [line.strip() for line in cell_text.split('\n') if line.strip()]
    
    subject = '-'
    lesson_type = '-'
    weeks = '-'
    teacher = '-'
    room = '-'
    
    # Обработаем первую строку, которая может содержать всё
    first_line = lines[0] if lines else ''
    
    # Ищем скобки с типом (Л/ПЗ/СЕМ и недели)
    lesson_match = re.search(r'\(([ЛПСлпс]+)[ПЗИМЕЛ]*\s*(\d+(?:-\d+)?(?:\D\s*\d+(?:-\d+)?)*)\s*н?\)', first_line)
    
    if lesson_match:
        # Есть скобки - разбиваем по ним
        lesson_type = lesson_match.group(1).upper()
        weeks = lesson_match.group(2).replace(' ', '')
        
        # Предмет = всё до скобок
        subject = first_line[:lesson_match.start()].strip()
        
        # Остальное после скобок = может быть преподаватель и аудитория
        remainder = first_line[lesson_match.end():].strip()
        if remainder:
            # Пробуем найти ФИО преподавателя
            teacher_match = re.search(r'([А-Я][а-я]+\s+[А-Я]\.[А-Я]\.)', remainder)
            if teacher_match:
                teacher = teacher_match.group(1)
                # Аудитория = остаток после преподавателя
                after_teacher = remainder[teacher_match.end():].strip()
                if after_teacher and re.match(r'^[А-ТУ-Я\w\-]', after_teacher):
                    room = after_teacher.split()[0]
            else:
                # Нет преподавателя, может быть только аудитория
                room_match = re.match(r'^([А-ТУ-Я\w\-]+\d+)', remainder)
                if room_match:
                    room = room_match.group(1)
    else:
        # Нет скобок в первой строке - это просто предмет
        subject = first_line
    
    # Обработаем остальные строки
    for line in lines[1:]:
        # Ищем информацию о типе/неделях если ещё не найдена
        if lesson_type == '-' or weeks == '-':
            lesson_match = re.search(r'\(([ЛПСлпс]+)[ПЗИМЕЛ]*\s*(\d+(?:-\d+)?(?:\D\s*\d+(?:-\d+)?)*)\s*н?\)', line)
            if lesson_match:
                lesson_type = lesson_match.group(1).upper()
                weeks = lesson_match.group(2).replace(' ', '')
        
        # Ищем преподавателя если не найден
        if teacher == '-':
            teacher_match = re.search(r'([А-Я][а-я]+\s+[А-Я]\.[А-Я]\.)', line)
            if teacher_match:
                teacher = teacher_match.group(1)
                continue
        
        # Ищем аудиторию если не найдена
        if room == '-':
            room_match = re.match(r'^([А-ТУ-Я\w\-]+\d+(?:/\d+)?(?:,\s*[А-ТУ-Я\w\-]+\d+)*)', line)
            if room_match:
                room = room_match.group(1)
                continue
    
    # Нормализуем название предмета
    subject = normalize_name(subject) if subject != '-' else '-'
    
    return subject, lesson_type, weeks, teacher, room





def parse_guu_schedule_all_sheets(excel_files: list) -> pd.DataFrame:
    """Парсер для всех листов в Excel файле."""
    
    print("🚀 Начинаю парсить ВСЕ листы расписания...\n")
    
    all_data = []
    
    time_slots = {
        '9.00-10.30': 1, '9.00-10,30': 1, '9,00-10,30': 1,
        '10.40-12.10': 2, '10.40-12,10': 2, '10,40-12,10': 2,
        '12.55-14.25': 3, '12.55-14,25': 3, '12,55-14,25': 3,
        '14.35-16.05': 4, '14.35-16,05': 4, '14,35-16,05': 4,
        '16.15-17.45': 5, '16.15-17,45': 5, '16,15-17,45': 5,
        '17.55-19.25': 6, '17.55-19,25': 6, '17,55-19,25': 6,
    }
    
    day_names = ['ПОНЕДЕЛЬНИК', 'ВТОРНИК', 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА', 'ВОСКРЕСЕНЬЕ']
    
    for excel_file in excel_files:
        if not os.path.exists(excel_file):
            print(f"⚠️  Файл не найден: {excel_file}")
            continue
        
        # Определяем курс, форму обучения и уровень образования из названия файла
        file_lower = excel_file.lower()
        
        if file_lower.startswith('m.'):
            # m.xlsx → Очная форма, Магистратура
            form = 'Очная'
            education_level = 'Магистратура'
            course = 1  # Для магистрантов курс обычно 1 или 2, но в названии листа будет указано
        elif file_lower.startswith('z.'):
            # z.xlsx → Очно-заочная форма, Бакалавриат
            form = 'Очно-заочная'
            education_level = 'Бакалавриат'
            course_match = re.search(r'(\d+)', excel_file)
            course = int(course_match.group(1)) if course_match else 1
        else:
            # 1.xlsx, 2.xlsx, 3.xlsx, 4.xlsx → Очная форма, Бакалавриат
            form = 'Очная'
            education_level = 'Бакалавриат'
            course_match = re.search(r'(\d+)', excel_file)
            course = int(course_match.group(1)) if course_match else 1
        
        print(f"📖 Парсирую: {excel_file} (Форма: {form}, Уровень: {education_level}, Курс: {course})")
        
        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception as e:
            print(f"❌ Ошибка чтения Excel: {e}")
            continue
        
        print(f"   Листов: {len(wb.sheetnames)}\n")
        
        # ===== ПАРСИМ КАЖДЫЙ ЛИСТ =====
        for sheet_name in wb.sheetnames:
            # Пропускаем пустые листы
            if sheet_name.lower() in ['лист1', 'лист', 'sheet1', 'sheet']:
                print(f"   ⏭️  Пропускаю пустой лист: {sheet_name}")
                continue
            
            # Определяем курс из названия листа (для магистратуры/очно-заочки)
            sheet_course = course
            sheet_form = form
            sheet_education_level = education_level
            
            if education_level == 'Магистратура' or form == 'Очно-заочная':
                # Ищем информацию о курсе в названии листа (например "1 к." или "2 курс")
                course_match = re.search(r'(\d+)\s*(?:к|курс)', sheet_name.lower())
                if course_match:
                    sheet_course = int(course_match.group(1))
            
            ws = wb[sheet_name]
            # Для магистратуры определим институт позже, если есть метаданные
            institute = get_institute_name(sheet_name)
            
            # Находим ключевые строки
            header_row = find_header_row(ws)
            time_row = find_time_row(ws, header_row)
            
            if header_row is None or time_row is None:
                print(f"   📚 Лист: '{sheet_name}' - ❌ Не найдена структура")
                continue
            
            print(f"   📚 Лист: '{sheet_name}'")
            print(f"      Институт: {institute}")
            
            # Получаем информацию о группах
            groups_dict = get_group_columns(ws, header_row)
            if not groups_dict:
                print(f"      ⚠️  Не найдены группы\n")
                continue
            
            # Получаем метаданные (направление и программу)
            is_magistracy = sheet_education_level == 'Магистратура'
            metadata_dict = get_group_metadata(ws, header_row, groups_dict, is_magistracy=is_magistracy, sheet_name=sheet_name)
            
            print(f"      ✅ Групп: {len(groups_dict)}")
            
            # Получаем информацию о позициях Дня, Времени, Недели
            day_col, time_col, week_col = get_day_time_columns(ws, time_row)
            
            if not all([day_col, time_col]):
                print(f"      ⚠️  Не найдены колонки День/Время\n")
                continue
            
            # ===== ПАРСИМ РАСПИСАНИЕ =====
            pair_count = 0
            current_day = None
            current_time = None  # Запомним время для использования в следующей строке
            
            for row_idx in range(time_row + 1, ws.max_row + 1):
                # Получаем день недели
                day_cell = ws.cell(row_idx, day_col)
                day_val = str(day_cell.value).upper().strip() if day_cell.value else ""
                
                if day_val in day_names:
                    current_day = day_val
                
                # Получаем время (может быть в этой строке)
                time_cell = ws.cell(row_idx, time_col)
                time_val = str(time_cell.value).strip() if time_cell.value else ""
                
                # Если есть время - запомним его (используется для этой И следующей строки)
                if time_val and time_val != '-':
                    current_time = time_val
                
                # Используем текущее время (из этой строки или сохранённое из предыдущей)
                time_to_use = current_time if current_time else time_val
                
                # === ВАЖНО: Обрабатываем ВСЕ строки с днём, даже если нет времени! ===
                # Пропускаем только если нет дня ИЛИ вообще нет времени было
                if not current_day or not time_to_use or time_to_use == '-':
                    continue
                
                # Определяем номер пары
                pair_num = time_slots.get(time_to_use, '-')
                
                # === ОПРЕДЕЛЯЕМ ЧЁТНОСТЬ ===
                # Приоритет:
                # 1. Если есть явное указание в week_col ("Чет." или "Нечет.") - используем его
                # 2. Если week_col пуст или нет - определяем по позиции строки
                
                base_parity = '-'
                
                # Сначала пробуем явное указание в week_col
                has_explicit_parity = False
                if week_col:
                    week_cell = ws.cell(row_idx, week_col)
                    week_val = str(week_cell.value).lower().strip() if week_cell.value else ""
                    
                    # Ищем явное указание четности 
                    # В excel написано "Чет." или "Нечет." или "Четная" или "Нечетная"
                    # ВНИМАНИЕ: В Excel'е ЧЕТ. означает НЕЧЁТНУЮ неделю (0), НЕЧЕТ. означает ЧЁТНУЮ неделю (1)!
                    if week_val.startswith('чет'):  # "чет.", "четная", "четн" -> это НЕЧЁТНАЯ (0)
                        base_parity = '0'
                        has_explicit_parity = True
                    elif week_val.startswith('нечет'):  # "нечет.", "нечетная", "нечетн" -> это ЧЁТНАЯ (1)
                        base_parity = '1'
                        has_explicit_parity = True
                
                # Если нет явного указания - определяем по позиции строки
                if not has_explicit_parity:
                    row_offset = row_idx - time_row  # Смещение от строки времени
                    if row_offset > 0:
                        if row_offset % 2 == 1:  # +1, +3, +5... = Чётная неделя (1)
                            base_parity = '1'
                        else:  # +2, +4, +6... = Нечётная неделя (0)
                            base_parity = '0'
                    else:
                        base_parity = '1'  # По умолчанию - Чётная
                
                # Обрабатываем каждую группу
                for col_idx, group_num in groups_dict.items():
                    cell = ws.cell(row_idx, col_idx)
                    cell_text = str(cell.value).strip() if cell.value else ""
                    
                    if not cell_text or cell_text == '-':
                        continue
                    
                    # Проверяем объединена ли ячейка!
                    cell_is_merged = is_cell_merged(ws, row_idx, col_idx)
                    
                    # ЛОКАЛЬНАЯ чётность для этой ячейки (1 = чётная, 0 = нечётная)
                    cell_parity = base_parity
                    
                    # Парсим содержимое ячейки
                    subject, lesson_type, weeks, teacher, room = parse_cell_content(cell_text)
                    
                    # Получаем направление и программу
                    meta = metadata_dict.get(col_idx, {'direction': '-', 'program': '-', 'institute': ''})
                    
                    # Если ячейка объединена (Обе недели), создаём ДВЕ записи!
                    parities_to_add = []
                    if cell_is_merged:
                        parities_to_add = ['1', '0']  # ДВЕ записи! (1=чётная, 0=нечётная)
                    else:
                        parities_to_add = [cell_parity]  # Одна запись
                    
                    # Добавляем запись(и)
                    for parity_value in parities_to_add:
                        # Определяем точный институт исходя из программы
                        # Для ОЗФО (Очно-заочная) - извлекаем из программы
                        # Для остального - используем обычную логику
                        if sheet_form == 'Очно-заочная':
                            exact_institute = extract_institute_from_program(meta['program'])
                            # Если программы нет или не распознана, пытаемся использовать направление
                            if not exact_institute:
                                exact_institute = extract_institute_from_program(meta['direction'])
                            # Если всё равно не нашли - используем пустое значение, НЕ sheet_name!
                            if not exact_institute:
                                exact_institute = ''
                        else:
                            # Передаём determined_institute из метаданных (из объединённых ячеек)
                            exact_institute = get_institute_name(sheet_name, meta['program'], meta['direction'], meta.get('institute', ''))
                        
                        record = {
                            'Форма обучения': sheet_form,
                            'Уровень образования': sheet_education_level,
                            'Курс': sheet_course,
                            'Институт': exact_institute,
                            'Направление': meta['direction'],
                            'Программа': meta['program'],
                            'Группа': f'Группа {group_num}',
                            'Номер группы': group_num,
                            'День недели': current_day,
                            'Номер пары': pair_num,
                            'Время пары': time_to_use,
                            'Чётность': parity_value,
                            'Предмет': subject,
                            'Вид пары': lesson_type,
                            'Преподаватель': teacher,
                            'Номер аудитории': room,
                            'Недели': weeks,
                        }
                        
                        all_data.append(record)
                        pair_count += 1
            
            print(f"      ✅ Пар: {pair_count}\n")
    
    # Создаём DataFrame
    if all_data:
        result_df = pd.DataFrame(all_data)
        column_order = [
            'Форма обучения', 'Уровень образования', 'Курс', 'Институт', 'Направление', 'Программа',
            'Группа', 'Номер группы', 'День недели', 'Номер пары', 'Время пары',
            'Чётность', 'Предмет', 'Вид пары', 'Преподаватель', 'Номер аудитории', 'Недели'
        ]
        result_df = result_df[column_order]
    else:
        result_df = pd.DataFrame()
    
    print(f"✅ Всего записей из всех листов: {len(result_df)}\n")
    
    return result_df


if __name__ == "__main__":
    excel_files = []
    for f in os.listdir('.'):
        # Ищем ВСЕ Excel файлы (1.xlsx, 2.xlsx, 3.xlsx, и т.д.)
        if f.endswith('.xlsx') and f not in ['schedule_full.xlsx']:
            excel_files.append(f)
    
    excel_files.sort()
    
    if not excel_files:
        print("❌ Не найдены файлы .xlsx расписания (1.xlsx, 2.xlsx, 3.xlsx и т.д.)")
        sys.exit(1)
    
    print(f"📂 Найдены файлы расписания: {', '.join(excel_files)}\n")
    df = parse_guu_schedule_all_sheets(excel_files)
    
    if len(df) > 0:
        # Сохраняем результаты
        output_csv = "schedule_full.csv"
        output_xlsx = "schedule_full.xlsx"
        
        df.to_csv(output_csv, index=False, encoding='utf-8')
        df.to_excel(output_xlsx, index=False, sheet_name='Расписание')
        
        print(f"💾 Сохранено:")
        print(f"   CSV:  {output_csv}")
        print(f"   XLSX: {output_xlsx}")
        
        print(f"\n📈 ФИНАЛЬНАЯ СТАТИСТИКА:")
        print(f"   Всего пар: {len(df)}")
        print(f"   Курсов: {df['Курс'].nunique()}")
        print(f"   Институтов: {df['Институт'].nunique()}")
        print(f"   Направлений: {len([x for x in df['Направление'].unique() if x != '-'])}")
        print(f"   Программ: {len([x for x in df['Программа'].unique() if x != '-'])}")
        print(f"   Групп: {df['Номер группы'].nunique()}")
        print(f"   Дней недели: {df['День недели'].nunique()}")
        print(f"   Предметов: {df['Предмет'].nunique()}")
        print(f"   Преподавателей: {df[df['Преподаватель'] != '-']['Преподаватель'].nunique()}")
        
        print(f"\n📚 КАЧЕСТВО ДАННЫХ:")
        print(f"   Преподаватели: {len(df[df['Преподаватель'] != '-'])}/{len(df)} = {100*len(df[df['Преподаватель'] != '-'])/len(df):.1f}%")
        print(f"   Аудитории: {len(df[df['Номер аудитории'] != '-'])}/{len(df)} = {100*len(df[df['Номер аудитории'] != '-'])/len(df):.1f}%")
        print(f"   Направления: {len(df[df['Направление'] != '-'])}/{len(df)} = {100*len(df[df['Направление'] != '-'])/len(df):.1f}%")
        print(f"   Программы: {len(df[df['Программа'] != '-'])}/{len(df)} = {100*len(df[df['Программа'] != '-'])/len(df):.1f}%")
    else:
        print("⚠️  Не удалось спарсить данные!")
